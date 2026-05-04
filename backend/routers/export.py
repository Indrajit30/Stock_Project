import io
import logging
import math
from datetime import datetime, timezone
from html import escape
from statistics import median

from fastapi import APIRouter, HTTPException, Request
from fastapi.encoders import jsonable_encoder
from fastapi.responses import Response
from reportlab.platypus import Paragraph

logger = logging.getLogger(__name__)
router = APIRouter()


def _cache(request: Request):
    return request.app.state.cache


async def _get_report(ticker: str, cache):
    if cache:
        data = await cache.get(f"stock:{ticker}:report:v2") or await cache.get(f"stock:{ticker}:report")
        if data:
            return data
    raise HTTPException(status_code=404, detail=f"No cached report for {ticker}. Trigger /api/stock/{ticker}/report first.")


def _as_float(value):
    try:
        if value is None:
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _money(value) -> str:
    n = _as_float(value)
    if n is None:
        return "N/A"
    if abs(n) >= 1_000_000_000_000:
        return f"${n / 1_000_000_000_000:.2f}T"
    if abs(n) >= 1_000_000_000:
        return f"${n / 1_000_000_000:.2f}B"
    if abs(n) >= 1_000_000:
        return f"${n / 1_000_000:.2f}M"
    return f"${n:,.0f}"


def _number(value, suffix: str = "") -> str:
    n = _as_float(value)
    if n is None:
        return "N/A"
    return f"{n:,.2f}{suffix}"


def _percent(value) -> str:
    n = _as_float(value)
    if n is None:
        return "N/A"
    if abs(n) <= 2:
        n *= 100
    return f"{n:.1f}%"


def _paragraph(text: object, style):
    return Paragraph(escape(str(text or "N/A")), style)


def _bullet_list(items: list, style, limit: int | None = None):
    selected = items[:limit] if limit else items
    if not selected:
        return [_paragraph("No data available.", style)]
    return [_paragraph(f"- {item}", style) for item in selected]


def _plain(value: object, fallback: str = "N/A") -> str:
    text = str(value or "").strip()
    return text if text else fallback


def _shorten(value: object, limit: int = 240) -> str:
    text = " ".join(_plain(value, "").split())
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "..."


def _clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def _safe_div(num, den):
    n = _as_float(num)
    d = _as_float(den)
    if n is None or d in (None, 0):
        return None
    return n / d


def _median(values: list[float | None]) -> float | None:
    clean = [v for v in values if v is not None and math.isfinite(v)]
    return median(clean) if clean else None


def _ratio_vs_peer(value, peer_value, lower_is_better: bool = True) -> str:
    v = _as_float(value)
    p = _as_float(peer_value)
    if v is None or p in (None, 0):
        return "Peer benchmark unavailable"
    diff = (v / p) - 1
    if abs(diff) < 0.1:
        return "Near peer median"
    cheaper = diff < 0 if lower_is_better else diff > 0
    direction = "discount" if cheaper else "premium"
    return f"{abs(diff) * 100:.0f}% {direction} to peer median"


def _metric_note(label: str, value: str, interpretation: str) -> list[str]:
    return [label, value, interpretation]


async def _get_export_data(ticker: str, cache) -> tuple[dict, dict]:
    from backend.routers.stock import REPORT_CACHE_VERSION, _api_key_configured, _fallback_report, _live_bundle

    report = None
    if cache:
        report = (
            await cache.get(f"stock:{ticker}:report:{REPORT_CACHE_VERSION}")
            or await cache.get(f"stock:{ticker}:report")
        )

    bundle = await _live_bundle(ticker, cache)

    if not report:
        report = _fallback_report(ticker, bundle)
        if cache:
            await cache.set(
                f"stock:{ticker}:report:{REPORT_CACHE_VERSION}",
                jsonable_encoder(report),
                3600,
            )
    elif not _api_key_configured():
        report.setdefault("hedging_detector", "OpenAI API key not configured; management-tone analysis skipped.")

    report.setdefault("ticker", ticker)
    report.setdefault("company_name", bundle.get("company_name") or ticker)
    report.setdefault("financials", bundle.get("financials") or {})
    report.setdefault("snowflake_scores", bundle.get("snowflake") or {})
    return report, bundle


async def _market_profile(ticker: str, cache) -> dict:
    from backend.services.data_fetcher import DataFetcher

    fetcher = DataFetcher(cache=cache)
    info = await fetcher._yfinance_info(ticker)
    profile = {
        "info": info or {},
        "major_holders": [],
        "institutional_holders": [],
        "recommendations_summary": [],
    }

    try:
        import asyncio
        import yfinance as yf

        def _records(frame):
            if frame is None:
                return []
            try:
                return frame.fillna("").head(12).to_dict("records")
            except Exception:
                return []

        def _fetch_optional():
            t = yf.Ticker(ticker)
            out = {}
            for key, getter in {
                "major_holders": t.get_major_holders,
                "institutional_holders": t.get_institutional_holders,
                "recommendations_summary": t.get_recommendations_summary,
            }.items():
                try:
                    out[key] = _records(getter())
                except Exception:
                    out[key] = []
            return out

        loop = asyncio.get_event_loop()
        optional = await loop.run_in_executor(None, _fetch_optional)
        profile.update(optional)
    except Exception as exc:
        logger.warning("Optional yfinance export data failed for %s: %s", ticker, exc)

    return profile


def _peer_candidates(ticker: str, fin: dict, info: dict) -> list[str]:
    sector = str(fin.get("sector") or info.get("sector") or "").lower()
    industry = str(fin.get("industry") or info.get("industry") or "").lower()
    text = f"{sector} {industry}"
    groups = [
        (("semiconductor", "chip"), ["AMD", "AVGO", "QCOM", "INTC", "TSM", "MU"]),
        (("software", "application", "infrastructure"), ["MSFT", "ORCL", "CRM", "ADBE", "NOW", "INTU"]),
        (("internet", "content", "advertising", "communication"), ["GOOGL", "META", "NFLX", "PINS", "SNAP"]),
        (("retail", "e-commerce", "discount"), ["AMZN", "WMT", "COST", "TGT", "HD"]),
        (("auto", "vehicle", "automaker"), ["TSLA", "F", "GM", "TM", "RIVN"]),
        (("bank", "financial"), ["JPM", "BAC", "WFC", "C", "GS", "MS"]),
        (("pharma", "biotech", "drug"), ["LLY", "JNJ", "PFE", "MRK", "ABBV", "BMY"]),
        (("consumer electronics", "hardware"), ["AAPL", "DELL", "HPQ", "STX", "WDC"]),
        (("energy", "oil", "gas"), ["XOM", "CVX", "COP", "EOG", "SLB"]),
    ]
    for keywords, candidates in groups:
        if any(keyword in text for keyword in keywords):
            return [symbol for symbol in candidates if symbol != ticker][:5]
    fallback = ["AAPL", "MSFT", "GOOGL", "AMZN", "META", "NVDA", "JPM", "V"]
    return [symbol for symbol in fallback if symbol != ticker][:5]


async def _peer_rows(ticker: str, fin: dict, profile: dict, cache) -> tuple[str, list[dict]]:
    from backend.services.data_fetcher import DataFetcher

    fetcher = DataFetcher(cache=cache)
    candidates = _peer_candidates(ticker, fin, profile.get("info") or {})
    batch = await fetcher.get_batch_financials([ticker, *candidates])
    rows = []
    for symbol in [ticker, *candidates]:
        snap = batch.get(symbol)
        if not snap:
            continue
        identity = await fetcher.get_company_identity(symbol)
        row = snap.model_dump() if hasattr(snap, "model_dump") else dict(snap)
        row["ticker"] = symbol
        row["company_name"] = identity.get("company_name") or symbol
        row["net_margin"] = _safe_div(row.get("net_income_ttm"), row.get("revenue_ttm"))
        rows.append(row)
    label = _plain(fin.get("industry") or profile.get("info", {}).get("industry") or fin.get("sector"), "Peer group")
    return label, rows


def _recommendation(fin: dict, profile: dict, peers: list[dict]) -> dict:
    info = profile.get("info") or {}
    growth = _as_float(fin.get("revenue_growth_yoy")) or _as_float(info.get("revenueGrowth"))
    pe = _as_float(fin.get("pe_ratio") or info.get("trailingPE"))
    forward_pe = _as_float(info.get("forwardPE"))
    ev_ebitda = _as_float(fin.get("ev_ebitda"))
    gross_margin = _as_float(fin.get("gross_margin"))
    net_margin = _safe_div(fin.get("net_income_ttm"), fin.get("revenue_ttm"))
    debt_to_equity = _as_float(fin.get("debt_to_equity"))
    fcf_yield = _safe_div(info.get("freeCashflow"), fin.get("market_cap") or info.get("marketCap"))
    peer_pe = _median([_as_float(row.get("pe_ratio")) for row in peers if row.get("ticker") != fin.get("ticker")])
    peer_margin = _median([_as_float(row.get("gross_margin")) for row in peers if row.get("ticker") != fin.get("ticker")])

    score = 0.0
    reasons = []
    risks = []
    if growth is not None:
        if growth >= 0.15:
            score += 2
            reasons.append(f"Revenue growth is strong at {_percent(growth)}.")
        elif growth >= 0.05:
            score += 1
            reasons.append(f"Revenue growth is positive at {_percent(growth)}.")
        elif growth < 0:
            score -= 2
            risks.append(f"Revenue is declining at {_percent(growth)}.")
        else:
            risks.append(f"Revenue growth is muted at {_percent(growth)}.")
    if gross_margin is not None:
        if peer_margin is not None and gross_margin > peer_margin:
            score += 1
            reasons.append(f"Gross margin of {_percent(gross_margin)} is above the peer median of {_percent(peer_margin)}.")
        elif gross_margin >= 0.5:
            score += 0.5
            reasons.append(f"Gross margin is high at {_percent(gross_margin)}.")
    if fcf_yield is not None:
        if fcf_yield > 0.03:
            score += 1
            reasons.append(f"Free-cash-flow yield is healthy at {_percent(fcf_yield)}.")
        elif fcf_yield < 0:
            score -= 1
            risks.append("Free cash flow is negative in the available market data.")
    if pe is not None and peer_pe is not None:
        if pe <= peer_pe * 0.9:
            score += 1
            reasons.append(f"P/E is below the peer median ({_number(pe)} vs {_number(peer_pe)}).")
        elif pe >= peer_pe * 1.3:
            score -= 1.5
            risks.append(f"P/E is materially above the peer median ({_number(pe)} vs {_number(peer_pe)}).")
    elif forward_pe is not None and forward_pe > 40:
        score -= 0.75
        risks.append(f"Forward P/E is elevated at {_number(forward_pe)}.")
    if debt_to_equity is not None:
        if debt_to_equity <= 1:
            score += 0.5
            reasons.append(f"Leverage appears manageable with debt/equity of {_number(debt_to_equity)}.")
        elif debt_to_equity > 3:
            score -= 1
            risks.append(f"Debt/equity is high at {_number(debt_to_equity)}.")

    if score >= 2:
        rating = "BUY"
    elif score <= -1.5:
        rating = "SELL"
    else:
        rating = "HOLD"

    confidence = _clamp(0.55 + abs(score) * 0.06, 0.55, 0.82)
    return {
        "rating": rating,
        "confidence": confidence,
        "score": score,
        "reasons": reasons[:4],
        "risks": risks[:4],
        "what_changes": _what_changes_rating(rating, fin, profile),
    }


def _what_changes_rating(rating: str, fin: dict, profile: dict) -> list[str]:
    info = profile.get("info") or {}
    growth = _as_float(fin.get("revenue_growth_yoy") or info.get("revenueGrowth"))
    target = _as_float(info.get("targetMeanPrice"))
    current = _as_float(info.get("currentPrice") or info.get("regularMarketPrice"))
    items = []
    if rating == "BUY":
        items.append("A downgrade would be warranted if growth slows sharply while the valuation multiple stays elevated.")
        items.append("A material balance-sheet deterioration or negative free-cash-flow trend would weaken the thesis.")
    elif rating == "SELL":
        items.append("An upgrade would require sustained revenue growth improvement plus a more reasonable valuation.")
        items.append("A stronger balance sheet and positive cash generation would reduce downside risk.")
    else:
        items.append("An upgrade requires clearer upside: accelerating growth, expanding margins, or a better entry price.")
        items.append("A downgrade would follow slowing demand, rising leverage, or valuation expanding without earnings support.")
    if growth is not None:
        items.append(f"Watch whether revenue growth can stay above {_percent(max(growth * 0.7, 0.05))}.")
    if current and target:
        items.append(f"Analyst target support weakens if price moves above the current mean target of {_money(target)}.")
    return items[:4]


def _forward_estimates(fin: dict, profile: dict) -> list[list[str]]:
    info = profile.get("info") or {}
    revenue = _as_float(fin.get("revenue_ttm") or info.get("totalRevenue"))
    eps = _as_float(info.get("trailingEps") or info.get("forwardEps"))
    growth = _as_float(fin.get("revenue_growth_yoy") or info.get("revenueGrowth"))
    earnings_growth = _as_float(info.get("earningsGrowth")) or growth
    base_growth = _clamp(growth if growth is not None else 0.05, -0.05, 0.25)
    bull_growth = _clamp(base_growth + 0.08, 0.02, 0.35)
    bear_growth = _clamp(base_growth - 0.10, -0.12, 0.12)
    base_eps_growth = _clamp(earnings_growth if earnings_growth is not None else base_growth, -0.08, 0.28)

    rows = [["Case", "Year 1 Revenue", "Year 2 Revenue", "Year 3 Revenue", "Year 3 EPS", "Assumption"]]
    for label, g, eps_g, note in [
        ("Bull", bull_growth, base_eps_growth + 0.05, "Demand and margins beat current trend"),
        ("Base", base_growth, base_eps_growth, "Current trend gradually normalizes"),
        ("Bear", bear_growth, base_eps_growth - 0.08, "Demand slows or valuation pressure rises"),
    ]:
        revs = []
        if revenue is not None:
            for year in range(1, 4):
                revs.append(_money(revenue * ((1 + g) ** year)))
        else:
            revs = ["N/A", "N/A", "N/A"]
        eps_year3 = _number(eps * ((1 + eps_g) ** 3)) if eps is not None else "N/A"
        rows.append([label, *revs, eps_year3, f"{_percent(g)} revenue CAGR; {note}"])
    return rows


def _valuation_rows(fin: dict, profile: dict, peers: list[dict]) -> tuple[list[list[str]], dict]:
    info = profile.get("info") or {}
    pe = _as_float(fin.get("pe_ratio") or info.get("trailingPE"))
    forward_pe = _as_float(info.get("forwardPE"))
    ev_ebitda = _as_float(fin.get("ev_ebitda"))
    growth = _as_float(fin.get("revenue_growth_yoy") or info.get("earningsGrowth") or info.get("revenueGrowth"))
    peg = pe / (growth * 100) if pe is not None and growth not in (None, 0) and growth > 0 else None
    peer_pe = _median([_as_float(row.get("pe_ratio")) for row in peers if row.get("ticker") != fin.get("ticker")])
    peer_ev = _median([_as_float(row.get("ev_ebitda")) for row in peers if row.get("ticker") != fin.get("ticker")])
    dcf = _simple_dcf(fin, profile)
    rows = [
        ["Metric", "Company", "Peer Median", "Interpretation"],
        ["Trailing P/E", _number(pe), _number(peer_pe), _ratio_vs_peer(pe, peer_pe)],
        ["Forward P/E", _number(forward_pe), "N/A", "Uses current market forward EPS if available"],
        ["PEG Ratio", _number(peg), "N/A", "Lower is better; high-growth names can support higher PEG"],
        ["EV/EBITDA", _number(ev_ebitda), _number(peer_ev), _ratio_vs_peer(ev_ebitda, peer_ev)],
        ["Simple DCF Value/Share", _money(dcf.get("value_per_share")), _money(info.get("currentPrice") or info.get("regularMarketPrice")), dcf.get("interpretation", "N/A")],
    ]
    return rows, dcf


def _simple_dcf(fin: dict, profile: dict) -> dict:
    info = profile.get("info") or {}
    fcf = _as_float(info.get("freeCashflow"))
    if fcf is None:
        fcf = (_as_float(fin.get("net_income_ttm")) or 0) * 0.8
    market_cap = _as_float(fin.get("market_cap") or info.get("marketCap"))
    current_price = _as_float(info.get("currentPrice") or info.get("regularMarketPrice"))
    shares = _as_float(info.get("sharesOutstanding"))
    if not shares and market_cap and current_price:
        shares = market_cap / current_price
    growth = _as_float(fin.get("revenue_growth_yoy") or info.get("revenueGrowth"))
    fcf_growth = _clamp(growth if growth is not None else 0.05, 0.02, 0.18)
    discount = 0.10
    terminal_growth = 0.03
    if fcf <= 0 or not shares:
        return {"value_per_share": None, "interpretation": "DCF unavailable because cash flow or share count is missing"}
    pv = 0
    for year in range(1, 6):
        future_fcf = fcf * ((1 + fcf_growth) ** year)
        pv += future_fcf / ((1 + discount) ** year)
    year5 = fcf * ((1 + fcf_growth) ** 5)
    terminal = year5 * (1 + terminal_growth) / (discount - terminal_growth)
    pv += terminal / ((1 + discount) ** 5)
    value_per_share = pv / shares
    if current_price:
        upside = (value_per_share / current_price) - 1
        interp = f"{_percent(upside)} implied upside/downside vs current price"
    else:
        interp = "Current price unavailable"
    return {
        "value_per_share": value_per_share,
        "fcf_growth": fcf_growth,
        "discount": discount,
        "terminal_growth": terminal_growth,
        "interpretation": interp,
    }


def _filing_insights(bundle: dict) -> list[str]:
    diff = bundle.get("filing_diff") or {}
    insights = []
    for section in (diff.get("changed_sections") or [])[:4]:
        name = _plain(section.get("section_name"), "Filing section")
        summary = _plain(section.get("summary"), "")
        positives = [item for item in (section.get("positives") or []) if len(str(item)) > 40][:2]
        negatives = [item for item in (section.get("negatives") or []) if len(str(item)) > 40][:1]
        if summary:
            insights.append(f"{name}: {summary}.")
        for item in positives:
            insights.append(f"{name} (positive): {_shorten(item, 220)}")
        for item in negatives:
            insights.append(f"{name} (negative): {_shorten(item, 220)}")
    return insights[:8]


def _risk_rows(fin: dict, profile: dict, bundle: dict) -> list[list[str]]:
    info = profile.get("info") or {}
    industry = str(fin.get("industry") or info.get("industry") or "").lower()
    filing_text = " ".join(_filing_insights(bundle)).lower()
    rows = [["Risk", "Why It Matters", "Downside Signal"]]
    rows.append(["Valuation compression", "High-multiple stocks can fall even when fundamentals improve.", f"P/E: {_number(fin.get('pe_ratio'))}; EV/EBITDA: {_number(fin.get('ev_ebitda'))}"])
    rows.append(["Growth slowdown", "The thesis weakens if revenue growth normalizes faster than margins or EPS can offset.", f"Current revenue growth: {_percent(fin.get('revenue_growth_yoy') or info.get('revenueGrowth'))}"])
    if "china" in filing_text or "export" in filing_text or "regulation" in filing_text:
        rows.append(["China / regulation exposure", "Filing language flags export controls or regulatory limits as relevant to demand and supply.", "Watch new export controls, licensing delays, and region-specific revenue disruption"])
    if "ai" in industry or "semiconductor" in industry or "software" in industry or "technology" in str(fin.get("sector") or "").lower():
        rows.append(["AI cycle risk", "AI infrastructure demand can be lumpy if customer capex pauses or product cycles slip.", "Watch order growth, inventory, and customer concentration"])
    debt = _as_float(fin.get("debt_to_equity"))
    if debt is not None and debt > 1:
        rows.append(["Balance-sheet risk", "Higher leverage reduces flexibility if earnings weaken.", f"Debt/equity: {_number(debt)}"])
    return rows[:6]


def _catalysts(fin: dict, profile: dict, bundle: dict) -> list[str]:
    info = profile.get("info") or {}
    sector = str(fin.get("sector") or info.get("sector") or "").lower()
    industry = str(fin.get("industry") or info.get("industry") or "").lower()
    items = [
        "Next earnings release: revenue growth, margin trend, and forward guidance.",
        "Estimate revisions or analyst target changes after new company disclosures.",
        "Free-cash-flow conversion and capital allocation updates.",
    ]
    if "semiconductor" in industry or "technology" in sector:
        items.insert(1, "Product-cycle execution and AI infrastructure demand updates.")
    if bundle.get("filing_diff", {}).get("changed_sections"):
        items.append("Resolution of key filing-change topics highlighted in the latest 10-Q comparison.")
    return items[:6]


def _capital_allocation_rows(fin: dict, profile: dict) -> list[list[str]]:
    info = profile.get("info") or {}
    total_cash = info.get("totalCash")
    total_debt = info.get("totalDebt")
    free_cash_flow = info.get("freeCashflow")
    dividend_yield = info.get("dividendYield")
    payout_ratio = info.get("payoutRatio")
    return [
        ["Metric", "Value", "Read-through"],
        ["Cash", _money(total_cash), "Dry powder for investment, buybacks, dividends, or debt reduction"],
        ["Debt", _money(total_debt), "Compare against cash generation and debt/equity"],
        ["Free Cash Flow", _money(free_cash_flow), "Primary fuel for capital returns and reinvestment"],
        ["Dividend Yield", _percent(dividend_yield), "Direct cash return to shareholders"],
        ["Payout Ratio", _percent(payout_ratio), "Dividend sustainability indicator"],
        ["Debt/Equity", _number(fin.get("debt_to_equity")), "Balance-sheet flexibility"],
    ]


def _key_metric_rows(fin: dict, profile: dict) -> list[list[str]]:
    info = profile.get("info") or {}
    revenue = _as_float(fin.get("revenue_ttm") or info.get("totalRevenue"))
    fcf = _as_float(info.get("freeCashflow"))
    net_income = _as_float(fin.get("net_income_ttm"))
    roe = _as_float(info.get("returnOnEquity"))
    roa = _as_float(info.get("returnOnAssets"))
    roic = _as_float(info.get("returnOnCapital")) or _as_float(info.get("returnOnCapitalEmployed"))
    fcf_margin = fcf / revenue if fcf is not None and revenue else None
    net_margin = net_income / revenue if net_income is not None and revenue else None
    return [
        ["Metric", "Value", "Investor Use"],
        ["ROIC", _percent(roic), "Measures return on capital deployed"],
        ["ROE", _percent(roe), "Measures return on shareholder equity"],
        ["ROA", _percent(roa), "Measures asset productivity"],
        ["FCF Margin", _percent(fcf_margin), "Shows cash conversion from revenue"],
        ["Net Margin", _percent(net_margin), "Shows bottom-line profitability"],
        ["Gross Margin", _percent(fin.get("gross_margin")), "Shows product/service profitability before opex"],
    ]


@router.get("/stock/{ticker}/export/pdf")
async def export_pdf(ticker: str, request: Request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    ticker = ticker.upper()
    cache = _cache(request)
    report, bundle = await _get_export_data(ticker, cache)
    profile = await _market_profile(ticker, cache)
    fin = dict(report.get("financials") or bundle.get("financials") or {})
    fin["ticker"] = ticker
    peer_label, peers = await _peer_rows(ticker, fin, profile, cache)
    thesis = _recommendation(fin, profile, peers)
    # Sync rating and confidence with the cached AI verdict so PDF matches the webpage
    _verdict_map = {"buy": "BUY", "wait": "HOLD", "avoid": "SELL"}
    if report.get("verdict"):
        thesis["rating"] = _verdict_map.get(str(report["verdict"]).lower(), thesis["rating"])
    if report.get("verdict_confidence") is not None:
        thesis["confidence"] = float(report["verdict_confidence"])
    valuation_rows, dcf = _valuation_rows(fin, profile, peers)
    info = profile.get("info") or {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=1.25*cm, rightMargin=1.25*cm, topMargin=1.25*cm, bottomMargin=1.25*cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="SectionTitle", parent=styles["Heading2"], textColor=colors.HexColor("#0f172a"), spaceBefore=14, spaceAfter=8))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.HexColor("#475569")))
    styles.add(ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=9, leading=11))
    styles.add(ParagraphStyle(name="Tiny", parent=styles["BodyText"], fontSize=7.2, leading=8.6, textColor=colors.HexColor("#334155")))
    styles.add(ParagraphStyle(name="CardTitle", parent=styles["Heading3"], fontSize=9.5, leading=11, textColor=colors.HexColor("#0f172a"), spaceAfter=2))
    styles.add(ParagraphStyle(name="Header", parent=styles["Title"], fontSize=16, leading=18, textColor=colors.HexColor("#0f172a")))
    styles.add(ParagraphStyle(name="HeaderBadge", parent=styles["BodyText"], fontSize=12, leading=14, alignment=1, textColor=colors.white))
    styles.add(ParagraphStyle(name="TableHeader", parent=styles["BodyText"], fontSize=7.4, leading=8.8, textColor=colors.white))
    story = []
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    company = report.get("company_name") or bundle.get("company_name") or ticker
    current_price = _as_float(info.get("currentPrice") or info.get("regularMarketPrice"))
    target_price = _as_float(info.get("targetMeanPrice"))

    def styled_table(data, col_widths, header="#0f172a"):
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEADING", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return table

    def para_rows(rows):
        return [
            [_paragraph(cell, styles["TableHeader"] if row_index == 0 else styles["Small"]) for cell in row]
            for row_index, row in enumerate(rows)
        ]

    rating = thesis["rating"]
    confidence = round(thesis["confidence"] * 100)
    rating_colors = {"BUY": colors.HexColor("#16a34a"), "HOLD": colors.HexColor("#f59e0b"), "SELL": colors.HexColor("#dc2626")}
    verdict_color = rating_colors.get(rating, colors.gray)
    peer_pe = _median([_as_float(row.get("pe_ratio")) for row in peers if row.get("ticker") != ticker])
    peer_ev = _median([_as_float(row.get("ev_ebitda")) for row in peers if row.get("ticker") != ticker])
    growth_rows = _forward_estimates(fin, profile)
    base_outlook = growth_rows[2] if len(growth_rows) > 2 else []

    entry_range = "N/A"
    target_range = "N/A"
    if current_price:
        if rating == "BUY":
            entry_range = f"{_money(current_price * 0.9)} - {_money(current_price * 1.03)}"
        elif rating == "HOLD":
            entry_range = f"Prefer pullbacks below {_money(current_price * 0.9)}"
        else:
            entry_range = f"Avoid new buys above {_money(current_price * 0.85)}"
    if target_price:
        target_range = f"{_money(target_price * 0.9)} - {_money(target_price * 1.1)}"
    elif dcf.get("value_per_share"):
        target_range = f"{_money(dcf['value_per_share'] * 0.9)} - {_money(dcf['value_per_share'] * 1.1)}"

    def card(title: str, flowables: list, accent="#0f172a"):
        title_para = Paragraph(f"<b>{escape(title)}</b>", styles["CardTitle"])
        content = [title_para, Spacer(1, 0.08*cm), *flowables]
        box = Table([[content]], colWidths=[8.75*cm])
        box.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#dbe3ef")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ("LINEABOVE", (0, 0), (-1, 0), 2.0, colors.HexColor(accent)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        return box

    def two_col(left, right):
        row = Table([[left, right]], colWidths=[8.95*cm, 8.95*cm])
        row.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        return row

    header_data = [
        [
            Paragraph(f"<b>{escape(company)} ({ticker})</b><br/><font size='8'>Generated {generated_at} | compact investment memo</font>", styles["Header"]),
            Paragraph(f"<b>{rating}</b><br/><font size='8'>{confidence}% confidence<br/>Price {_money(current_price)} | Target {_money(target_price)}</font>", styles["HeaderBadge"]),
        ]
    ]
    header = Table(header_data, colWidths=[12.2*cm, 5.7*cm])
    header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#f8fafc")),
        ("BACKGROUND", (1, 0), (1, 0), verdict_color),
        ("TEXTCOLOR", (1, 0), (1, 0), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(header)
    story.append(Spacer(1, 0.18*cm))

    thesis_bullets = (thesis["reasons"] or ["Available data supports a neutral stance until growth, valuation, or margin evidence improves."])[:3]
    thesis_bullets.append(thesis["what_changes"][0])
    thesis_card = card(
        "1. Investment Thesis",
        _bullet_list(thesis_bullets, styles["Tiny"], limit=4),
        "#1e3a8a",
    )

    metrics_rows = [
        ["Metric", "Value"],
        ["Revenue", _money(fin.get("revenue_ttm"))],
        ["Net Income", _money(fin.get("net_income_ttm"))],
        ["Growth", _percent(fin.get("revenue_growth_yoy") or info.get("revenueGrowth"))],
        ["Margins", f"Gross {_percent(fin.get('gross_margin'))}; Net {_percent(_safe_div(fin.get('net_income_ttm'), fin.get('revenue_ttm')))}"],
        ["P/E", _number(fin.get("pe_ratio") or info.get("trailingPE"))],
    ]
    metrics_card = card(
        "2. Key Metrics Snapshot",
        [styled_table(para_rows(metrics_rows), [3.7*cm, 4.55*cm], "#0f766e")],
        "#0f766e",
    )
    story.append(two_col(thesis_card, metrics_card))

    growth_lines = [
        f"Current revenue growth is {_percent(fin.get('revenue_growth_yoy') or info.get('revenueGrowth'))}.",
        f"Base case points to Year 1 revenue of {base_outlook[1] if len(base_outlook) > 1 else 'N/A'} and Year 3 revenue of {base_outlook[3] if len(base_outlook) > 3 else 'N/A'}.",
        "Main drivers are demand trend, pricing/margins, product cycle execution, and guidance revisions.",
        "Forecasts are model-derived, not company guidance.",
    ]
    growth_card = card(
        "3. Growth & Outlook",
        _bullet_list(growth_lines, styles["Tiny"], limit=4),
        "#065f46",
    )

    valuation_summary = [
        ["Metric", "Company", "Peer"],
        ["P/E", _number(fin.get("pe_ratio") or info.get("trailingPE")), _number(peer_pe)],
        ["EV/EBITDA", _number(fin.get("ev_ebitda")), _number(peer_ev)],
        ["DCF/share", _money(dcf.get("value_per_share")), _money(current_price)],
    ]
    valuation_text = _ratio_vs_peer(fin.get("pe_ratio") or info.get("trailingPE"), peer_pe)
    valuation_card = card(
        "4. Valuation",
        [
            styled_table(para_rows(valuation_summary), [2.8*cm, 2.7*cm, 2.7*cm], "#7c2d12"),
            Spacer(1, 0.08*cm),
            _paragraph(f"Read-through: {valuation_text}.", styles["Tiny"]),
        ],
        "#7c2d12",
    )
    story.append(two_col(growth_card, valuation_card))

    risk_items = [
        f"{row[0]}: {row[2]}"
        for row in _risk_rows(fin, profile, bundle)[1:5]
    ]
    risks_card = card(
        "5. Risks",
        _bullet_list(risk_items, styles["Tiny"], limit=4),
        "#991b1b",
    )

    filing_items = _filing_insights(bundle)[:4]
    if not filing_items:
        filing_items = ["No material filing-change summary was available for this ticker."]
    filing_card = card(
        "Filing / Catalyst Notes",
        _bullet_list([_shorten(item, 135) for item in filing_items[:2]] + _catalysts(fin, profile, bundle)[:2], styles["Tiny"], limit=4),
        "#334155",
    )
    story.append(two_col(risks_card, filing_card))

    analyst_rows = [
        ["Item", "Value"],
        ["Analyst view", _plain(info.get("recommendationKey") or info.get("recommendationMean"))],
        ["Opinions", _plain(info.get("numberOfAnalystOpinions"))],
        ["Inst. own.", _percent(info.get("heldPercentInstitutions"))],
        ["Insider own.", _percent(info.get("heldPercentInsiders"))],
    ]
    analyst_card = card(
        "Market Context",
        [styled_table(para_rows(analyst_rows), [3.2*cm, 5.0*cm], "#581c87")],
        "#581c87",
    )

    final_rows = [
        ["Recommendation", rating],
        ["Entry Range", entry_range],
        ["Target Range", target_range],
        ["Why", _shorten("; ".join(thesis["reasons"][:2]) or "Risk/reward is balanced.", 150)],
        ["Changes If", _shorten(thesis["what_changes"][0], 150)],
    ]
    final_card = card(
        "6. Final Recommendation",
        [styled_table(para_rows([["Item", "Conclusion"], *final_rows]), [3.2*cm, 5.0*cm], "#111827")],
        "#111827",
    )
    story.append(two_col(analyst_card, final_card))

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("<i>Data sourced from live fundamentals, SEC EDGAR-derived filing data, yfinance market data, and local valuation calculations. Estimates are model-derived and informational only; this is not financial advice.</i>", styles["Tiny"]))

    doc.build(story)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={ticker}_report.pdf"},
    )


@router.get("/stock/{ticker}/export/excel")
async def export_excel(ticker: str, request: Request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    ticker = ticker.upper()
    cache = _cache(request)
    report = await _get_report(ticker, cache)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1A1A2E")
    alt_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")

    def _style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill

    def _auto_width(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # Sheet 1 — Summary
    ws1 = wb.active
    ws1.title = "Summary"
    fin = report.get("financials", {})
    sf = report.get("snowflake_scores", {})
    rows = [
        ["Field", "Value"],
        ["Ticker", report.get("ticker")],
        ["Company", report.get("company_name")],
        ["Verdict", report.get("verdict")],
        ["Confidence", report.get("verdict_confidence")],
        ["Summary", report.get("plain_english_summary")],
        ["Revenue TTM", fin.get("revenue_ttm")],
        ["Net Income TTM", fin.get("net_income_ttm")],
        ["PE Ratio", fin.get("pe_ratio")],
        ["EV/EBITDA", fin.get("ev_ebitda")],
        ["Gross Margin", fin.get("gross_margin")],
        ["Debt/Equity", fin.get("debt_to_equity")],
        ["Market Cap", fin.get("market_cap")],
        ["Snowflake Value", sf.get("value")],
        ["Snowflake Growth", sf.get("growth")],
        ["Snowflake Health", sf.get("health")],
        ["Snowflake Momentum", sf.get("momentum")],
        ["Snowflake Smart Money", sf.get("smart_money")],
    ]
    for i, row in enumerate(rows):
        ws1.append(row)
        if i > 0 and i % 2 == 0:
            for cell in ws1[i + 1]:
                cell.fill = alt_fill
    _style_header(ws1)
    _auto_width(ws1)

    # Sheet 2 — Financials (placeholder — needs quarterly data)
    ws2 = wb.create_sheet("Financials")
    ws2.append(["Quarter", "Revenue", "Net Income", "Gross Margin"])
    _style_header(ws2)
    _auto_width(ws2)

    # Sheet 3 — Peer Comparison
    ws3 = wb.create_sheet("Peer Comparison")
    ws3.append(["Ticker", "Company", "Market Cap", "PE", "EV/EBITDA", "Gross Margin", "D/E"])
    if cache:
        peers = await cache.get(f"peers:{ticker}:list")
        if peers:
            for p in peers:
                ws3.append([p.get("ticker"), p.get("company_name"), p.get("market_cap"),
                             p.get("pe_ratio"), p.get("ev_ebitda"), p.get("gross_margin"), p.get("debt_to_equity")])
    _style_header(ws3)
    _auto_width(ws3)

    # Sheet 4 — Filing Diff
    ws4 = wb.create_sheet("Filing Diff")
    ws4.append(["Section", "Type", "Text"])
    if cache:
        diff = await cache.get(f"stock:{ticker}:filing_diff")
        if diff:
            for sec in diff.get("changed_sections", []):
                for p in sec.get("positives", []):
                    ws4.append([sec["section_name"], "Positive", p])
                for n in sec.get("negatives", []):
                    ws4.append([sec["section_name"], "Negative", n])
    _style_header(ws4)
    _auto_width(ws4)

    # Sheet 5 — Insider Trades
    ws5 = wb.create_sheet("Insider Trades")
    ws5.append(["Name", "Role", "Shares", "Value USD", "Trade Date", "10b5 Plan"])
    if cache:
        cluster = await cache.get(f"stock:{ticker}:insider_cluster")
        if cluster:
            for t in cluster.get("insiders", []):
                ws5.append([t.get("name"), t.get("role"), t.get("shares"),
                             t.get("value_usd"), t.get("trade_date"), t.get("is_10b5_plan")])
    _style_header(ws5)
    _auto_width(ws5)

    # Sheet 6 — Congressional Trades
    ws6 = wb.create_sheet("Congressional Trades")
    ws6.append(["Politician", "Chamber", "Party", "Transaction", "Amount", "Date"])
    if cache:
        cong = await cache.get(f"stock:{ticker}:congressional_trades")
        if cong:
            for t in cong:
                ws6.append([t.get("politician_name"), t.get("chamber"), t.get("party"),
                             t.get("transaction_type"), t.get("amount_range"), t.get("trade_date")])
    _style_header(ws6)
    _auto_width(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={ticker}_report.xlsx"},
    )
